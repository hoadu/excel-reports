import openpyxl

class ReportReader:
    def read(self, input_file):
        wb = openpyxl.load_workbook(input_file)
        data_sheet = wb.get_sheet_by_name('Sheet1')
        max_row = data_sheet.max_row
        max_column = data_sheet.max_column
        all_columns = data_sheet.columns